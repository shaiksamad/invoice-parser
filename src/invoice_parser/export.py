import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
# from  openpyxl.styles.differential import DifferentialStyle

from table import Table


class Export:
    """
    Takes input table and export to the Excel

    Parameters
    ----------
    table: Table
        list of rows to print

    """
    def __init__(self, table: Table):
        self.book = Workbook()
        self.sheet = self.book.active
        self._num_empty_rows = 0
        self.table = table
        self.table.sort_by_invoice()

        self.make_sheet()
        self.make_sheet('gold', rows=[*table.filter_by_item('gold')])
        self.make_sheet('silver', rows=[*table.filter_by_item('silver')])

        self.book.remove(self.book.get_sheet_by_name('Sheet'))

    @property
    def empty_row(self):
        self._num_empty_rows += 1
        return []

    @property
    def num_empty_rows(self):
        return self._num_empty_rows

    def save(self, filename: str) -> None:
        self.book.save(filename)

    def _set_formats(self, sheet=None):
        if not sheet:
            sheet = self.sheet

        ALIGN_CENTER = Alignment("center", "center", wrap_text=True, shrink_to_fit=True)
        DATE_FORMAT = '[$-en-US]dd-mmm-yy;@'
        SIDE = Side(None, "000000", "thin")
        BORDER_TOP_BOTTOM = Border(top=SIDE, bottom=SIDE)
        BORDER_ALL = Border(SIDE, SIDE, SIDE, SIDE)
        FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        NUMBER_FORMAT_2 = '0.00'
        NUMBER_FORMAT_3 = '0.000'
        for row in range(sheet.min_row, sheet.max_row + 1):
            for col in range(sheet.min_column, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                if row != sheet.max_row and col <= 9: # sheet.max_column - 2:
                    cell.border = BORDER_ALL

                if row == sheet.min_row:  # header
                    cell.alignment = ALIGN_CENTER
                    # cell.font = Font(bold=True)

                elif row == sheet.max_row:  # footer
                    if cell.value:
                        cell.border = BORDER_TOP_BOTTOM
                        cell.fill = FILL_YELLOW
                        sheet.row_dimensions[row].height = 18
                        cell.alignment = Alignment(vertical='center')
                        cell.number_format = NUMBER_FORMAT_2 if col != sheet.min_column + 3 else NUMBER_FORMAT_3
                else:
                    if isinstance(cell.value, (int, float)) and col > sheet.min_column:
                        cell.number_format = NUMBER_FORMAT_2 if col != sheet.min_column + 3 else NUMBER_FORMAT_3
                        sheet.column_dimensions[get_column_letter(col)].width = 10
                    elif isinstance(cell.value, datetime.datetime):
                        cell.number_format = DATE_FORMAT
                        cell.alignment = ALIGN_CENTER
                        sheet.column_dimensions[get_column_letter(col)].width = 12
                    elif isinstance(cell.value, str):
                        cell.alignment = ALIGN_CENTER
                        sheet.column_dimensions[get_column_letter(col)].width = 8

        max_sheet_col = get_column_letter(sheet.max_column + 2)
        sheet.conditional_formatting.add(
            f'{max_sheet_col}{sheet.min_row+3}:{max_sheet_col}{sheet.max_row+1}',
            CellIsRule('=', ['FALSE'], True, None, None, FILL_YELLOW)
        )

    def shift(self, rows: int = 2, cols: int = 2, sheet=None):
        sheet = sheet or self.sheet
        sheet.move_range(sheet.dimensions, rows=rows, cols=cols, translate=True)

    def make_sheet(self, name: str = "main", rows: list = None):
        sheet = self.book.create_sheet(name)
        table = self.table
        rows = rows or table.rows

        sheet.append(table.header)
        table.row_start = 2
        table.row_end = table.row_start + len(rows) - 1

        for row in rows:
            sheet.append(row)

        if name != 'main':
            table.footer[3] = f'=SUM(D{table.row_start}:D{table.row_end})'
        sheet.append(table.footer)

        self._set_formats(sheet=sheet)
        self.shift(sheet=sheet)


# todo: optimize export set_formats and document it
#         and do other todos
